# El path de mi configuracion de POWERHSELL es:
# C:\Users\jcrea\Powershell\Microsoft.PowerShell_profile.ps1

import os
import sys
import pandas as pd
import subprocess
import shutil
import PyPDF2
from datetime import datetime
from openpyxl import load_workbook
import time
import winsound
import re

start_time = time.time()

# -------------------- VARIABLES --------------------
template = "Template.xlsx" # plantilla
destination_file = "000IndiceExpedienteElectronico.xlsx" # indice
parts_file = "v_historico_oct_2023.csv" # Archivo con la informacion de las partes

working_directory = os.getcwd()
script_location = os.path.dirname(os.path.abspath(__file__))

class bcolors:
	HEADER = '\033[95m'
	OKBLUE = '\033[94m'
	OKCYAN = '\033[96m'
	OKGREEN = '\033[92m'
	WARNING = '\033[93m'
	FAIL = '\033[91m'
	ENDC = '\033[0m'
	BOLD = '\033[1m'
	UNDERLINE = '\033[4m'

# get comand line arguments
update_only = len(sys.argv) > 1 and sys.argv[1] == "update"
print("args")
print(sys.argv)
print(bcolors.OKGREEN + "[INFO - MODE] Modo de actualización: " + str(update_only))

# TODO: Feat partes A, B

def convert_bytes(num):
	"""
	Convert file size in bytes to a human-readable format.
	:param num: File size in bytes.
	:return: A string representing the file size in a human-readable format.
	"""
	for unit in ['bytes', 'KB', 'MB', 'GB', 'TB']:
		if num < 1024.0:
			return f"{num:3.1f} {unit}"
		num /= 1024.0

print("Trabajando en el directorio: " + working_directory)
print("Script ubicado en: " + script_location)

print("cargando archivo de partes...")
print(os.path.join(script_location, parts_file))

df = pd.read_csv(os.path.join(script_location, parts_file), sep=",")

parts_errors = []
				
# name of the own script file
script_file = os.path.basename(__file__)

def get_pdf_page_count(file_path):
	"""
	Get the number of pages of a PDF file.
	"""
	pdf = PyPDF2.PdfReader(open(file_path, "rb"))
	return len(pdf.pages)
	
def get_docx_page_count(file_path):
	output = "_output.pdf"
	try:
		subprocess.run(['docto', '-f', file_path, '-O', output, '-T', 'wdFormatPDF'])
	except Exception as e:
		print(bcolors.WARNING + "DOCTO error")
		print(bcolors.WARNING + "ERROR: No se pudo convertir el archivo: " + file_path)

	# Open the PDF file
	open_file = open(output, 'rb')
	pdf = PyPDF2.PdfReader(open_file)
	pages = len(pdf.pages)

	# close the PDF file
	open_file.close()
	print(bcolors.OKCYAN + f"\t[DOCX] The document {file_path} has {pages} pages.")

	# remove output file
	os.remove(output)
	return pages

def get_file_metadata(root, file, consecutivo):
	"""
	Get metadata of a file.
	"""
	file_path = os.path.normpath(os.path.join(root, file))
	metadata = {
		"File Name": file,
		"Size (Bytes)": os.path.getsize(file_path),
		"Last Modified": datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S'),
		"Created": datetime.fromtimestamp(os.path.getctime(file_path)).strftime('%Y-%m-%d %H:%M:%S'),
		"File format": os.path.splitext(file)[1],
	}
	
	page_count = 1

	if metadata["File format"] in [".pdf", ".PDF"]:
		try:
			page_count = get_pdf_page_count(file_path)
		except Exception as e:
			page_count = 0
		
	if metadata["File format"] in [".docx", ".DOCX"]:
		page_count = get_docx_page_count(file_path)

	base_structure = {
		"Nombre Documento": {
			"col": 1,	
			"data": file 
		},
		"Fecha Creacion Documento": {
			"col": 2,	
			"data": datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%d-%m-%Y')
		},
		"Fecha incorporacion expediente": {
			"col": 3,	
			"data": datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%d-%m-%Y'),
		},
		"Orden documento": {
			"col": 4,	
			"data": consecutivo
		},
		"Numero paginas": {
			"col": 5,	
			"data": page_count
		},
		"Formato": {
			"col": 8,	
			"data": os.path.splitext(file)[1]
		},
		"Tamanio": {
			"col": 9,	
			"data": convert_bytes(os.path.getsize(file_path)),
		},
		"Origen": {
			"col": 10,	
			"data": "Electrónico"
		},
	}

	return base_structure

def get_path_components(path):
	components = []

	while True:
		path, component = os.path.split(path)
		if component == "":
			break
		components.append(component)

	components.reverse()
	
	return components

def get_radicado_from_path(path):
	"""
	Clean the path of a file.
	"""
	# take last folder from path
	path_components = get_path_components(os.path.normpath(path))
	radicado = None

	for component in path_components:
		if component.isdigit():
			radicado = component
			break
	return radicado

def check_if_matches_substr(reference_str, substr_list):
	# NOTE: path clean up only for windows
	reference_str = get_radicado_from_path(reference_str)
	
	if not reference_str:
		return False
	
	for substr in substr_list:
		if reference_str in substr:
			return True
	return False

def main():
	# Traverse the current directory and its subdirectories
	# root = carpeta actual
	# files = lista de archivos en la carpeta actual
	# . = carpeta actual
	#  /imagenes
	#  /imagenes/imagen1.jpg
	#  /imagenes/imagen2.jpg
	#  /imagenes/imagen3.jpg
	#  /imagenes/paseo_santa_marta/imagen4.jpg
	
	folders_to_check = get_folders_to_check()
	checked_folders_list = None

	# TODO: Check proper skip of folders
	if folders_to_check:
		checked_folders_list = {}
		for folder in folders_to_check:
			checked_folders_list[folder] = False

	base_root = None
	conflict_filepaths = []
	success_processed_files = 0

	for root, _, files in os.walk("."):
		if not base_root:
			base_root = root

		if folders_to_check:
			if check_if_matches_substr(root, folders_to_check): 
				pass
			else:
				continue
		
		if checked_folders_list:
			checked_folders_list[get_radicado_from_path(root)] = True
			print(bcolors.OKGREEN + "> [MODE: Filtro por carpeta] Procesando carpeta: {}".format(root))

		# Initialize an empty list to store metadata dictionaries for each folder
		folder_files_metadata = []
		
		if len(files) == 0:
			continue

		archivo_copia = os.path.join(root, destination_file)
		print(bcolors.OKGREEN + "[INFO] Copiando archivo: {}".format(archivo_copia))
		shutil.copy(os.path.join(script_location, template), archivo_copia)
		
		consecutivo = 1

		for file in files:
			if (file == destination_file) or (file == template):
				continue

			try:
				metadata = get_file_metadata(root, file, consecutivo)
				consecutivo += 1
				folder_files_metadata.append(metadata)
			except Exception as e:
				print(bcolors.WARNING + "ERROR en el archivo: " + file)
				print(bcolors.WARNING + "ERROR en la carpeta: " + root)
				print(e)
				print("-" * 50)
				winsound.MessageBeep(-1)
				conflict_filepaths.append(os.path.join(root, file))

		# Skip empty folders
		if not folder_files_metadata:
			continue

		# Load the existing Excel workbook
		workbook = load_workbook(filename=archivo_copia)

		# Select the worksheet you want to modify (assuming the first sheet)
		worksheet = workbook.active

		fila_inicio = 10

		try:
			print(bcolors.OKGREEN + "[CARPETA] Procesando archivos en la carpeta: {}".format(root))
			id = extract_id_proceso_from_filename(root)

			if id:
				full = get_parts_a_b_data(id)
				
				if full:
					a, b, clase = full
					if a and b:
						if clase:
							worksheet.cell(row=4, column=2, value=clase)
						worksheet.cell(row=5, column=2, value=id)
						worksheet.cell(row=6, column=2, value=a)
						worksheet.cell(row=7, column=2, value=b)
					print("\t Escribiendo partes de proceso: " + id)

			for index, file_metadata in enumerate(folder_files_metadata):
				print(bcolors.OKBLUE + "\t[ARCHIVO] Procesando archivo: {}".format(file_metadata["Nombre Documento"]["data"]))
				success_processed_files += 1

				for key, val in file_metadata.items():
					worksheet.cell(row=fila_inicio + index, column=val["col"], value=val["data"])
		except Exception as e:
			print(bcolors.WARNING + "ERROR en la fila: " + str(fila_inicio + index))
			print(bcolors.WARNING + "ERROR en el archivo: " + file)
			
			winsound.MessageBeep(-1)
			# rethrow error
			raise e

		# Save the changes
		workbook.save(filename=os.path.join(root, destination_file))
	
	if len(conflict_filepaths) != 0:
		print(bcolors.WARNING + "[ADVERTENCIA] Los siguientes archivos no pudieron ser procesados:\n")
		print(bcolors.WARNING + "\n".join(conflict_filepaths))
		print()
	else:
		print(bcolors.OKGREEN + "[INFO] Todos los archivos fueron procesados correctamente.")

	print(bcolors.OKGREEN + "[INFO] Se procesaron {} archivos.".format(success_processed_files))
	
	if parts_errors and not update_only:
		print(bcolors.WARNING + "[ADVERTENCIA] No fue posible llenar la informacion de las partes para los siguientes radicados")
		# print unique parts_errors
		print(bcolors.WARNING + "\n".join(list(set(parts_errors))))

	if checked_folders_list:
		print(bcolors.WARNING + "\n[ADVERTENCIA] Las siguientes carpetas no fueron procesadas:")
		for folder, checked in checked_folders_list.items():
			if not checked:
				print("\t" + bcolors.WARNING + folder)

def check_large_filepaths():
	"""
	Check if the length of the file path is too long.
	"""
	long_file_paths = []
	
	for dirpath, _, files in os.walk("."):
		# Saltar carpetas sin archivos
		if len(files) == 0:
			continue

		for file in files:
			file_path = os.path.join(dirpath, file)
			if len(file_path) > 255:
				long_file_paths.append(file_path)
				
	if len(long_file_paths) != 0:
		print(bcolors.WARNING + "[ADVERTENCIA] Los siguientes archivos son muy largos:")
		print(bcolors.WARNING + "\n".join(long_file_paths))
	else:
		print(bcolors.OKGREEN + "No se encontraron archivos con rutas muy largas.")

	if input("\n¿Deseas continuar? [s/n]: ").lower() == "n":
		exit()
		
def extract_id_proceso_from_filename(filename):
	if filename == ".":
		filename = os.getcwd()
	
	print("[INFO] Extrayendo id de proceso del nombre de la carpeta: " + filename)

	pattern = r"\\(\d+)(\\)?"
	match = re.search(pattern, filename)
	
	if match:
		return match.group(1)
	else:
		return None
		
def get_folders_to_check():
	# The user can enter a list of folders to check
	folders_to_check = input("Ingresa las carpetas a revisar (separadas por coma): ") 
	folders_to_check = list(map(lambda x: x.strip(), folders_to_check.split(",")))
	
	if not folders_to_check:
		print(bcolors.WARNING + "No se ingresaron carpetas a revisar.")
		confirmation = input("¿Deseas revisar todas las carpetas? [s/n]: ")
		
		if confirmation.lower() == "s":
			return None
		exit()
		
	if folders_to_check[0] == '':
		return None

	return folders_to_check

def get_parts_a_b_data(id_proceso):
    # query the df and output "demandado", "demandante", and "clase" where id_proceso == id_proceso
    header_proceso = "radicado"
    try:
        a = df.loc[df[header_proceso] == id_proceso]['demandado'].values[0]
        b = df.loc[df[header_proceso] == id_proceso]['demandante'].values[0]

        print("\t[INFO] Se encontró el proceso con id: " + id_proceso)
        print("\t[INFO] Parte A: " + a)
        print("\t[INFO] Parte B: " + b)

        # Check if header 'clase' exists
        if 'clase' in df.columns:
            clase = df.loc[df[header_proceso] == id_proceso]['clase'].values[0]
            print("\t[INFO] Clase: " + clase)
        else:
            print("\t[INFO] No Hay Clase en el archivo de partes")

        return [a, b, clase]
    except Exception as e:
        parts_errors.append(id_proceso)
        print(bcolors.WARNING + "ERROR: No se encontró el proceso con id: " + id_proceso)
        return None
	
def similarity(word1, word2):
	"""
	Calculates the similarity between two words using the Levenshtein distance algorithm.
	Returns a value between 0 and 1, where 0 means the words are completely different and 1 means they are identical.
	"""
	if len(word1) < len(word2):
		return similarity(word2, word1)

	if len(word2) == 0:
		return 1.0

	previous_row = range(len(word2) + 1)
	for i, c1 in enumerate(word1):
		current_row = [i + 1]
		for j, c2 in enumerate(word2):
			insertions = previous_row[j + 1] + 1
			deletions = current_row[j] + 1
			substitutions = previous_row[j] + (c1 != c2)
			current_row.append(min(insertions, deletions, substitutions))
		previous_row = current_row

	return 1.0 - (previous_row[-1] / max(len(word1), len(word2)))

# traverse all files in the current directory
def check_nonunified_files():
	conflict_files = []
	
	for dirpath, _, files in os.walk("."):
		# check if there is a file with .xlsx extension
		for file in files:
			if file.endswith(".xlsx") and file != destination_file and similarity(file.lower(), destination_file.lower()) >= 0.3:
				conflict_files.append(os.path.join(dirpath, file))
				try:
					os.rename(os.path.join(dirpath, file), os.path.join(dirpath, destination_file))
				except Exception as e:
					# delete the file
					try:
						os.remove(os.path.join(dirpath, file))
					except Exception as e:
						print(bcolors.WARNING + "ERROR: No se pudo eliminar el archivo: " + os.path.join(dirpath, file))
				
	# print(bcolors.WARNING + "[ADVERTENCIA] Los siguientes no estan unificados:")
	# print(bcolors.WARNING + "\n".join(conflict_files))
	
main()

end_time = time.time()
elapsed_time = end_time - start_time

print(bcolors.OKGREEN + "Tiempo de ejecución: " + str(elapsed_time) + " segundos")
